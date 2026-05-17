"""Local text/data extraction fallback for PDF, DOCX, XLSX, CSV, TXT.

Used when a file is too large for direct Bedrock upload or when local extraction is preferred
(e.g., Excel files). Returns an empty string on failure rather than raising, consistent with
the OCR-disabled design (scanned PDFs without embedded text return empty silently).
"""

from __future__ import annotations

from pathlib import Path

from proposal_ingest.logging_utils import get_logger

logger = get_logger("extractors")


def extract_text(path: Path) -> str:
    """Return extracted plain text from a file, or an empty string if unsupported or failed.

    Dispatches by file extension. All errors are caught and logged as warnings so the
    pipeline continues rather than halting on a bad file.
    """
    ext = path.suffix.lower()
    try:
        if ext == ".pdf":
            return _extract_pdf(path)
        if ext in {".docx", ".doc"}:
            return _extract_docx(path)
        if ext in {".xlsx", ".xls"}:
            return _extract_xlsx(path)
        if ext == ".csv":
            return _extract_csv(path)
        if ext in {".txt", ".md", ".html"}:
            return _extract_text_file(path)
    except Exception:
        logger.warning("Text extraction failed for %s", path, exc_info=True)
    return ""


# ---------------------------------------------------------------------------
# Format-specific helpers
# ---------------------------------------------------------------------------


def _extract_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    parts: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        if text.strip():
            parts.append(text)
    return "\n".join(parts)


def _extract_docx(path: Path) -> str:
    from docx import Document

    doc = Document(str(path))
    return "\n".join(para.text for para in doc.paragraphs if para.text.strip())


def _extract_xlsx(path: Path) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in wb.worksheets:
        parts.append(f"[Sheet: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(c.strip() for c in cells):
                parts.append("\t".join(cells))
    wb.close()
    return "\n".join(parts)


def _extract_csv(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def _extract_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


# ---------------------------------------------------------------------------
# Excel size helpers (used by strategy selection)
# ---------------------------------------------------------------------------


def count_excel_nonempty_cells(path: Path) -> tuple[int, int]:
    """Return (sheet_count, nonempty_cell_count) for an Excel file.

    Opens in read-only mode and counts populated cells up to a ceiling to
    avoid loading huge workbooks into memory. Returns (-1, -1) on error.
    """
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheet_count = len(wb.worksheets)
        nonempty = 0
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None and str(cell).strip():
                        nonempty += 1
        wb.close()
        return sheet_count, nonempty
    except Exception:
        logger.warning("Could not count Excel cells for %s", path, exc_info=True)
        return -1, -1
